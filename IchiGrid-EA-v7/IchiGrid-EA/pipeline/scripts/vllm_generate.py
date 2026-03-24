#!/usr/bin/env python3
"""
vLLM Batch Code Generator for IchiGridEA
==========================================
Reads EA_Master_vLLM.xlsx → dispatches modules to 20 AI workers → generates code.

Usage:
  python vllm_generate.py --xlsx EA_Master_vLLM.xlsx --url http://localhost:8000/v1
  python vllm_generate.py --xlsx EA_Master_vLLM.xlsx --url http://localhost:8000/v1 --workers 10 --model deepseek-coder-v2

Prerequisites:
  pip install openai openpyxl

Workflow:
  1. In Excel, set Status = "ready" for modules you want generated
  2. Run this script → 20 workers generate code in parallel
  3. Generated files go to ./generated/ folder
  4. Excel is updated: Status, AssignedTo, CodeHash, GenDate, GenModel, Notes
"""

import argparse
import hashlib
import os
import sys
import time
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    from openai import OpenAI
except ImportError:
    print("ERROR: openai required. Run: pip install openai")
    input("Press Enter to exit...")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Run: pip install openpyxl")
    input("Press Enter to exit...")
    sys.exit(1)


# ── System prompt (shared by all 20 workers) ──
SYSTEM_PROMPT = """You are an expert MQL5 developer for the IchiGridEA algorithmic trading system.
You generate production-quality MetaTrader 5 code.
Rules:
- One class per .mqh file, with #ifndef guards
- Include @module and @hash tags in file headers
- All public methods must have error handling and logging
- Use strict MQL5 typing (no implicit conversions)
- Follow IchiGridEA module isolation architecture
- Code must compile without warnings in MT5 Build 4540+
"""


def load_modules(xlsx_path, status_filter='ready'):
    """Load modules from Excel where Status = status_filter."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['PLAN']
    modules = []
    for row in range(3, ws.max_row + 1):
        sid = ws.cell(row=row, column=1).value
        status = (ws.cell(row=row, column=12).value or '').strip().lower()
        if not sid or status != status_filter:
            continue
        modules.append({
            'row': row,
            'Section': str(sid),
            'Path': ws.cell(row=row, column=2).value or '',
            'ModuleName': ws.cell(row=row, column=3).value or '',
            'Files': ws.cell(row=row, column=4).value or '',
            'Extensions': ws.cell(row=row, column=5).value or '',
            'Function': ws.cell(row=row, column=6).value or '',
            'Dependencies': ws.cell(row=row, column=7).value or '',
            'Category': ws.cell(row=row, column=8).value or '',
            'PlanHash': ws.cell(row=row, column=10).value or '',
        })
    return modules


def build_prompt(mod):
    """Build the code generation prompt for a single module."""
    return f"""Generate the complete MQL5 code for this module:

## Module Specification
- Section: {mod['Section']}
- Path: {mod['Path']}
- Module: {mod['ModuleName']}
- Files to generate: {mod['Files']}
- Function: {mod['Function']}
- Dependencies: {mod['Dependencies']}
- PlanHash: {mod['PlanHash']}

## Requirements
1. Generate ALL files listed above
2. .mqh: #ifndef guards, class with constructor/destructor
3. .json: default schema with documented fields
4. .csv: column headers matching the module's data model
5. File header: // @module: {mod['ModuleName']} | @hash: {mod['PlanHash']} | @gen: {datetime.now().strftime('%Y-%m-%d')}
6. #include all Dependencies: {mod['Dependencies']}
7. Error handling + logging in every public method
8. Wrap in: // gen:begin:{mod['ModuleName']} ... // gen:end:{mod['ModuleName']}

## Output
Return ONLY code. One code block per file, with filename as comment. No explanations."""


def generate_one(client, model, mod, output_dir):
    """Generate code for a single module via vLLM API."""
    try:
        resp = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": build_prompt(mod)}
            ],
            max_tokens=8192,
            temperature=0.2,
        )
        code = resp.choices[0].message.content
        code_hash = hashlib.sha256(code.encode()).hexdigest()

        # Save code to file
        safe_name = ''.join(c if c.isalnum() or c == '_' else '_' for c in mod['ModuleName'][:60])
        fpath = os.path.join(output_dir, f"{mod['Section']}_{safe_name}.mqh")
        with open(fpath, 'w', encoding='utf-8') as f:
            f.write(code)

        return {
            'status': 'done',
            'code_hash': code_hash,
            'model': model,
            'file': fpath,
            'tokens': getattr(resp.usage, 'completion_tokens', 0) if resp.usage else 0,
        }
    except Exception as e:
        return {
            'status': 'error',
            'code_hash': '',
            'model': model,
            'file': '',
            'error': str(e)[:200],
        }


def worker(worker_id, modules, vllm_url, model, output_dir):
    """One AI worker processes its assigned modules."""
    client = OpenAI(base_url=vllm_url, api_key="none")
    my_mods = [m for m in modules if m.get('AssignedTo') == f'IA-{worker_id:02d}']
    results = []
    for mod in my_mods:
        print(f"  [IA-{worker_id:02d}] [{mod['Section']:>5s}] {mod['ModuleName'][:45]}...", end=' ')
        t0 = time.time()
        result = generate_one(client, model, mod, output_dir)
        elapsed = time.time() - t0
        status_icon = '✅' if result['status'] == 'done' else '❌'
        print(f"{status_icon} ({elapsed:.1f}s)")
        results.append((mod, result))
        time.sleep(0.1)  # Rate limiting
    return results


def update_excel(xlsx_path, all_results):
    """Update Excel with generation results."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb['PLAN']
    ws_hash = wb['HASHMAP'] if 'HASHMAP' in wb.sheetnames else None
    now_str = datetime.now().strftime('%Y-%m-%d')

    for mod, result in all_results:
        row = mod['row']
        ws.cell(row=row, column=12).value = result['status']         # Status
        ws.cell(row=row, column=14).value = mod.get('AssignedTo','') # AssignedTo
        code_hash = result.get('code_hash', '')
        ws.cell(row=row, column=15).value = code_hash[:16] + '...' if code_hash else ''  # CodeHash
        ws.cell(row=row, column=16).value = now_str                   # GenDate
        ws.cell(row=row, column=17).value = result.get('model', '')   # GenModel
        if result.get('error'):
            ws.cell(row=row, column=18).value = result['error']       # Notes

        # Update HASHMAP full hash
        if ws_hash and code_hash:
            for hr in range(2, ws_hash.max_row + 1):
                if str(ws_hash.cell(row=hr, column=1).value) == mod['Section']:
                    ws_hash.cell(row=hr, column=3).value = code_hash
                    ws_hash.cell(row=hr, column=4).value = now_str
                    break

    wb.save(xlsx_path)


def main():
    parser = argparse.ArgumentParser(description='vLLM Batch Code Generator for IchiGridEA')
    parser.add_argument('--xlsx', required=True, help='Path to EA_Master_vLLM.xlsx')
    parser.add_argument('--url', default='http://localhost:8000/v1', help='vLLM API URL')
    parser.add_argument('--model', default='codellama/CodeLlama-34b-Instruct-hf', help='Model name')
    parser.add_argument('--workers', type=int, default=20, help='Number of parallel AI workers (1-20)')
    parser.add_argument('--output-dir', default='./generated', help='Output directory for generated code')
    parser.add_argument('--status', default='ready', help='Process modules with this status (default: ready)')
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)

    print(f"\n{'='*70}")
    print(f"vLLM BATCH CODE GENERATOR — IchiGridEA")
    print(f"{'='*70}")
    print(f"  Excel:   {args.xlsx}")
    print(f"  vLLM:    {args.url}")
    print(f"  Model:   {args.model}")
    print(f"  Workers: {args.workers}")
    print(f"  Output:  {args.output_dir}")
    print(f"  Filter:  Status = '{args.status}'")

    # Load modules
    modules = load_modules(args.xlsx, args.status)
    if not modules:
        print(f"\n  No modules with Status='{args.status}'.")
        print(f"  Open {args.xlsx}, set Status='ready' for modules to generate.")
        input("\nPress Enter to exit...")
        return

    print(f"\n  Found {len(modules)} modules to generate.")

    # Assign round-robin to workers
    n_workers = min(args.workers, len(modules))
    for i, mod in enumerate(modules):
        mod['AssignedTo'] = f'IA-{(i % n_workers) + 1:02d}'

    # Show assignment summary
    from collections import Counter
    assignments = Counter(m['AssignedTo'] for m in modules)
    print(f"\n  Worker assignments:")
    for w, cnt in sorted(assignments.items()):
        print(f"    {w}: {cnt} modules")

    print(f"\n{'─'*70}")
    print(f"Starting generation...\n")

    t0 = time.time()
    all_results = []
    with ThreadPoolExecutor(max_workers=n_workers) as executor:
        futures = [
            executor.submit(worker, w+1, modules, args.url, args.model, args.output_dir)
            for w in range(n_workers)
        ]
        for future in as_completed(futures):
            all_results.extend(future.result())

    elapsed = time.time() - t0
    done = sum(1 for _, r in all_results if r['status'] == 'done')
    errors = sum(1 for _, r in all_results if r['status'] == 'error')

    print(f"\n{'─'*70}")
    print(f"RESULTS:")
    print(f"  ✅ Done:   {done}")
    print(f"  ❌ Errors: {errors}")
    print(f"  ⏱  Time:   {elapsed:.1f}s ({elapsed/len(modules):.1f}s/module)")

    # Update Excel
    print(f"\nUpdating {args.xlsx}...")
    update_excel(args.xlsx, all_results)
    print(f"✅ Excel updated with Status, CodeHash, GenDate, GenModel.")

    # Error summary
    if errors:
        print(f"\n⚠️  Error details:")
        for mod, result in all_results:
            if result['status'] == 'error':
                print(f"  [{mod['Section']}] {mod['ModuleName'][:40]}: {result.get('error','')[:80]}")

    input("\nPress Enter to exit...")


if __name__ == '__main__':
    main()
