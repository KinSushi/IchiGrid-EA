#!/usr/bin/env python3
"""
excel_hash_sync.py v2 — Synchronise les hashes LLM + disque vers Excel + HASHES.json
======================================================================================

DEUX COUCHES DE HASH:
  1. @code_hash (gen:header)  = Ce que le LLM a inscrit a la generation -> Excel CodeHash (col15)
  2. SHA-256 fichier disque   = Integrite actuelle du fichier           -> HASHES.json + Notes

Le script extrait TOUTES les metadonnees du gen:header:
  @code_hash, @spec_hash, @model, @generated, @session, @section, @version, @depends

Usage:
  python excel_hash_sync.py --repo /path/to/IchiGrid-EA \
                            --excel /path/to/EA_Master_vLLM.xlsx \
                            --output /path/to/output.xlsx \
                            --update-hashes --report SYNC_REPORT.md

Version: 2.0.0 | Date: 2026-03-23
"""

import argparse, hashlib, json, os, re, sys
from datetime import datetime, timezone

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    sys.exit("ERREUR: pip install openpyxl")

# ── CONSTANTES ──
TRACKED_EXT = {'.mqh','.mq5','.json','.csv','.set','.txt','.md','.png','.log','.zip','.dat','.py','.editorconfig'}
IGNORED_DIRS = {'.git','__pycache__','node_modules','.venv'}
MQL5_EXT = {'.mqh','.mq5'}
COL = dict(SECTION=1,PATH=2,MODULE=3,FILES=4,EXT=5,FUNC=6,DEPS=7,CAT=8,TAGS=9,
           PLAN_HASH=10,SRC=11,STATUS=12,PRIO=13,ASSIGNED=14,CODE_HASH=15,DATE=16,MODEL=17,NOTES=18)
HM = dict(SECTION=1,PLAN_HASH=2,CODE_HASH=3,VERIFIED=4)
FIRST_ROW = 3

# ── REGEX gen:header ──
RX = {
    'code_hash':  re.compile(r'@code_hash\s*:\s*([a-fA-F0-9]+)'),
    'spec_hash':  re.compile(r'@spec_hash\s*:\s*([a-fA-F0-9]+)'),
    'model':      re.compile(r'@model\s*:\s*(\S+)'),
    'generated':  re.compile(r'@generated(?:_at)?\s*:\s*(\S+)'),
    'session':    re.compile(r'@session(?:_id)?\s*:\s*(\S+)'),
    'section':    re.compile(r'@section\s*:\s*(\S+)'),
    'version':    re.compile(r'@version\s*:\s*(\S+)'),
    'module':     re.compile(r'@module\s*:\s*(\S+)'),
}

def sha256_file(p):
    h = hashlib.sha256()
    with open(p,'rb') as f:
        for c in iter(lambda: f.read(65536), b''): h.update(c)
    return h.hexdigest()

def sha256_str(s): return hashlib.sha256(s.encode()).hexdigest()

def composite_hash(hh):
    if len(hh)==1: return hh[0]
    return sha256_str('|'.join(sorted(hh)))

def count_lines(p):
    try:
        with open(p,'r',encoding='utf-8',errors='ignore') as f: return sum(1 for _ in f)
    except: return 0

def parse_files(s):
    if not s or s.strip() in ('None','','nan'): return []
    return [x.strip() for x in s.split(',') if x.strip() and x.strip()!='None']

def parse_gen_header(content):
    """Extrait toutes les metadonnees du gen:header (100 premieres lignes)."""
    zone = '\n'.join(content.split('\n')[:100])
    out = {}
    for key, rx in RX.items():
        m = rx.search(zone)
        if m:
            v = m.group(1).strip()
            if not (v.startswith('{') and v.endswith('}')):  # skip placeholders
                out[key] = v
    return out

# ── SCAN REPO ──
def scan_repo(repo):
    index = {}  # basename -> [entry, ...]
    entries = []
    total_bytes = 0
    for root, dirs, files in os.walk(repo):
        dirs[:] = [d for d in dirs if d not in IGNORED_DIRS]
        for fname in files:
            ext = os.path.splitext(fname)[1].lower()
            if ext not in TRACKED_EXT and fname not in ('LICENSE','.gitignore'): continue
            ap = os.path.join(root, fname)
            rp = os.path.relpath(ap, repo)
            sz = os.path.getsize(ap)
            dh = sha256_file(ap)
            ln = count_lines(ap)
            gh = {}
            if ext in MQL5_EXT:
                try:
                    with open(ap,'r',encoding='utf-8',errors='ignore') as f: gh = parse_gen_header(f.read())
                except: pass
            e = {'rel':rp,'abs':ap,'base':fname,'ext':ext,'sha256':dh,'size':sz,'lines':ln,'gh':gh}
            index.setdefault(fname,[]).append(e)
            entries.append(e)
            total_bytes += sz

    mql = [e for e in entries if e['ext'] in MQL5_EXT]
    gh_stats = {
        'total_mql5': len(mql),
        'with_code_hash': sum(1 for e in mql if 'code_hash' in e['gh']),
        'with_spec_hash': sum(1 for e in mql if 'spec_hash' in e['gh']),
        'with_model': sum(1 for e in mql if 'model' in e['gh']),
        'with_session': sum(1 for e in mql if 'session' in e['gh']),
        'with_section': sum(1 for e in mql if 'section' in e['gh']),
    }
    return {'index':index,'entries':entries,'total_files':len(entries),'total_bytes':total_bytes,'gh_stats':gh_stats}

def pick_best(candidates):
    for c in candidates:
        if c['rel'].startswith('Include/'): return c
    for c in candidates:
        if c['rel'].startswith('Core/'): return c
    return candidates[0]

def match_section(expected, idx):
    matched, missing = {}, []
    for f in expected:
        if f in idx: matched[f] = pick_best(idx[f])
        else: missing.append(f)
    if not matched:
        return {'status':'none','matched':{},'missing':missing,'code_hash':None,'disk_hash':None,
                'gen_model':None,'gen_date':None,'drift':False,'drift_files':[],
                'expected':len(expected),'found':0}

    llm_hh, disk_hh, drifts = {}, {}, []
    for f in sorted(matched):
        e = matched[f]
        disk_hh[f] = e['sha256']
        gh_ch = e['gh'].get('code_hash')
        if gh_ch:
            llm_hh[f] = gh_ch
            if not e['sha256'].startswith(gh_ch): drifts.append(f)
        else:
            llm_hh[f] = e['sha256']  # fallback

    code_hash = composite_hash(list(llm_hh.values()))
    disk_hash = composite_hash(list(disk_hh.values()))

    gen_model, gen_date = None, None
    for f in sorted(matched):
        gh = matched[f]['gh']
        if 'model' in gh:
            gen_model = gh['model']
            gen_date = gh.get('generated')
            break

    return {'status':'full' if not missing else 'partial',
            'matched':matched,'missing':missing,
            'code_hash':code_hash,'disk_hash':disk_hash,
            'gen_model':gen_model,'gen_date':gen_date,
            'drift':len(drifts)>0,'drift_files':drifts,
            'expected':len(expected),'found':len(matched)}

# ── SYNC EXCEL ──
def sync_excel(excel_path, output_path, repo_idx, now_str):
    wb = openpyxl.load_workbook(excel_path)
    wp = wb['PLAN']; wh = wb['HASHMAP']
    st = dict(total=0,full=0,partial=0,none=0,written=0,updated=0,
              llm=0,disk=0,drifts=0,by_ext={},files=set(),sec_by_file={})

    hm_idx = {}
    for r in range(2, wh.max_row+1):
        s = str(wh.cell(r,HM['SECTION']).value or '').strip()
        if s: hm_idx[s] = r

    gf, of = Font(color="006400"), Font(color="CC6600")

    for r in range(FIRST_ROW, wp.max_row+1):
        sec = str(wp.cell(r,COL['SECTION']).value or '').strip()
        if not sec: continue
        st['total'] += 1
        fs = str(wp.cell(r,COL['FILES']).value or '')
        cur_status = str(wp.cell(r,COL['STATUS']).value or '').strip()
        expected = parse_files(fs)
        if not expected: st['none']+=1; continue
        for ef in expected:
            ext = os.path.splitext(ef)[1].lower()
            st['by_ext'][ext] = st['by_ext'].get(ext,0)+1

        res = match_section(expected, repo_idx)
        if res['status']=='full': st['full']+=1
        elif res['status']=='partial': st['partial']+=1
        else: st['none']+=1; continue

        if not res['code_hash']: continue

        # CodeHash = @code_hash LLM (priorite)
        wp.cell(r,COL['CODE_HASH']).value = res['code_hash']
        st['written'] += 1

        if res['gen_model']:
            wp.cell(r,COL['MODEL']).value = res['gen_model']
            st['llm'] += 1
        else:
            wp.cell(r,COL['MODEL']).value = "sha256-sync/v2"
            st['disk'] += 1

        wp.cell(r,COL['DATE']).value = res['gen_date'] or now_str

        if res['status']=='full' and cur_status in ('spec','draft','','None'):
            wp.cell(r,COL['STATUS']).value = 'done'
            wp.cell(r,COL['STATUS']).font = gf
            st['updated'] += 1
        elif res['status']=='partial' and cur_status in ('spec','draft','','None'):
            wp.cell(r,COL['STATUS']).value = 'review'
            wp.cell(r,COL['STATUS']).font = of
            st['updated'] += 1

        # Notes: disk hash + drift + missing
        parts = []
        if res['disk_hash']: parts.append(f"disk:{res['disk_hash'][:16]}")
        if res['drift']:
            parts.append(f"DRIFT:{','.join(res['drift_files'][:3])}")
            st['drifts'] += 1
        if res['missing']: parts.append(f"MISSING:{','.join(res['missing'][:3])}")

        old = str(wp.cell(r,COL['NOTES']).value or '')
        cleaned = re.sub(r'(disk:[a-f0-9]+|DRIFT:\S+|MISSING:\S+|PARTIAL:\S+)','',old).strip('; ')
        auto = ' | '.join(parts)
        final = f"{cleaned}; {auto}" if cleaned and cleaned!='None' else auto
        wp.cell(r,COL['NOTES']).value = final[:500] if final else None

        for f in res['matched']:
            st['files'].add(f)
            st['sec_by_file'].setdefault(f,[]).append(sec)
        if sec in hm_idx:
            wh.cell(hm_idx[sec],HM['CODE_HASH']).value = res['code_hash']
            wh.cell(hm_idx[sec],HM['VERIFIED']).value = now_str

    # Update STATS
    if 'STATS' in wb.sheetnames:
        ws = wb['STATS']
        sc = {}
        for rr in range(FIRST_ROW, wp.max_row+1):
            s = str(wp.cell(rr,COL['STATUS']).value or '').strip()
            if s: sc[s] = sc.get(s,0)+1
        for r in range(1,ws.max_row+1):
            lb = str(ws.cell(r,1).value or '').strip()
            if lb=='Code Generated': ws.cell(r,2).value = st['full']
            elif lb=='In Review': ws.cell(r,2).value = st['partial']
            elif lb=='Done': ws.cell(r,2).value = st['full']
            elif lb=='Remaining': ws.cell(r,2).value = st['none']
            elif lb in sc: ws.cell(r,2).value = sc[lb]

    wb.save(output_path)
    st['files'] = sorted(st['files'])
    st['sec_by_file'] = {k:len(v) for k,v in st['sec_by_file'].items()}
    return st

# ── SYNC HASHES.json ──
def sync_hashes_json(hp, excel_path, scan, now_str):
    with open(hp) as f: hd = json.load(f)

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['PLAN']
    phm = {}
    for r in range(FIRST_ROW, ws.max_row+1):
        sec = str(ws.cell(r,COL['SECTION']).value or '').strip()
        ph = str(ws.cell(r,COL['PLAN_HASH']).value or '').strip()
        fs = str(ws.cell(r,COL['FILES']).value or '')
        if sec and ph and ph!='None':
            for fn in parse_files(fs):
                phm.setdefault(fn,[]).append({'section':sec,'plan_hash':ph})

    nf = {}; tl = 0
    for e in scan['entries']:
        ext = e['ext']
        if ext not in ('.mqh','.mq5','.json','.csv','.set','.txt','.md','.png','.log','.dat','.zip'): continue
        if e['rel'].startswith('pipeline/') or e['rel'].startswith('tests/python/'): continue

        old = hd.get('files',{}).get(e['rel'],{})
        section = old.get('section','')
        spec_hash = ''

        gh = e['gh']
        if gh.get('section') and gh['section'] not in ('{SECTION_ID}','UNKNOWN'):
            section = gh['section']
        if e['base'] in phm:
            ep = phm[e['base']]
            if not section and ep: section = ep[0]['section']
            aph = sorted(set(x['plan_hash'] for x in ep))
            spec_hash = composite_hash(aph) if len(aph)>1 else aph[0]
        if gh.get('spec_hash'): spec_hash = gh['spec_hash']

        rec = {'sha256':e['sha256'],'lines':e['lines'],'section':section,'spec_hash':spec_hash}
        if gh.get('code_hash'): rec['gen_code_hash'] = gh['code_hash']
        if gh.get('model'): rec['gen_model'] = gh['model']
        if gh.get('generated'): rec['gen_date'] = gh['generated']
        if gh.get('session'): rec['gen_session'] = gh['session']
        if gh.get('version'): rec['gen_version'] = gh['version']
        if gh.get('code_hash') and not e['sha256'].startswith(gh['code_hash']):
            rec['drift'] = True

        nf[e['rel']] = rec
        tl += e['lines']

    hd['files'] = nf
    hd['_generated'] = now_str
    hd['_version'] = '7.0'
    hd['_files'] = len(nf)
    hd['_lines'] = tl

    with open(hp,'w',encoding='utf-8') as f: json.dump(hd, f, indent=2, ensure_ascii=False)
    return {
        'total':len(nf),'lines':tl,
        'spec_hash':sum(1 for v in nf.values() if v.get('spec_hash')),
        'gen_code_hash':sum(1 for v in nf.values() if v.get('gen_code_hash')),
        'sections':sum(1 for v in nf.values() if v.get('section')),
        'drifts':sum(1 for v in nf.values() if v.get('drift')),
    }

# ── RAPPORT ──
def report(es, hs, sc, now):
    gh = sc['gh_stats']
    L = [
        f"# Rapport sync v2 — {now}\n",
        f"## Repo: {sc['total_files']} fichiers, {sc['total_bytes']/1024:.1f} KB\n",
        "## gen:header LLM",
        f"- MQL5: **{gh['total_mql5']}** | @code_hash: **{gh['with_code_hash']}** | @model: **{gh['with_model']}** | @section: **{gh['with_section']}**\n",
        "## Excel PLAN",
        f"- Total: **{es['total']}** | done: **{es['full']}** | review: **{es['partial']}** | pending: **{es['none']}**",
        f"- CodeHash ecrits: **{es['written']}** (LLM: {es['llm']}, disk fallback: {es['disk']})",
        f"- Drifts: **{es['drifts']}** (fichier modifie apres generation)\n",
        "## Fichiers matches",
    ]
    for f,c in sorted(es.get('sec_by_file',{}).items(), key=lambda x:-x[1]):
        L.append(f"- `{f}` -> {c} sections")
    L.append("\n## Extensions attendues")
    for ext,c in sorted(es.get('by_ext',{}).items(), key=lambda x:-x[1]):
        L.append(f"- `{ext}`: {c}")
    if hs:
        L.extend([
            "\n## HASHES.json",
            f"- Fichiers: **{hs['total']}** | gen_code_hash: **{hs['gen_code_hash']}** | spec_hash: **{hs['spec_hash']}** | drifts: **{hs['drifts']}**",
        ])
    L.extend([
        "\n## Legende Notes (col18)",
        "- `disk:abc123...` = SHA-256 integrite fichier disque",
        "- `DRIFT:File.mqh` = @code_hash LLM != SHA-256 disque",
        "- `MISSING:File.json` = fichier attendu absent du repo",
        f"\n---\n`excel_hash_sync.py v2` — {now}",
    ])
    return '\n'.join(L)

# ── MAIN ──
def main():
    p = argparse.ArgumentParser(description='Sync hashes LLM+disque -> Excel+HASHES.json v2')
    p.add_argument('--repo', required=True)
    p.add_argument('--excel', required=True)
    p.add_argument('--output', required=True)
    p.add_argument('--update-hashes', action='store_true')
    p.add_argument('--report', default=None)
    p.add_argument('--dry-run', action='store_true')
    a = p.parse_args()

    now = datetime.now(timezone.utc)
    ns = now.strftime('%Y-%m-%dT%H:%M:%SZ')
    nd = now.strftime('%Y-%m-%d')

    print(f"[1/5] Scan: {a.repo}")
    sc = scan_repo(a.repo)
    gh = sc['gh_stats']
    print(f"  {sc['total_files']} fichiers | MQL5: {gh['total_mql5']} | @code_hash: {gh['with_code_hash']} | @model: {gh['with_model']}")

    if not a.dry_run:
        print(f"[2/5] Excel -> {a.output}")
        es = sync_excel(a.excel, a.output, sc['index'], nd)
    else:
        print("[2/5] DRY RUN"); es = dict(total=0,full=0,partial=0,none=0,written=0,updated=0,llm=0,disk=0,drifts=0,by_ext={},files=[],sec_by_file={})
    print(f"  done:{es['full']} review:{es['partial']} pending:{es['none']} | CodeHash:{es['written']} (LLM:{es['llm']} disk:{es['disk']}) | drifts:{es['drifts']}")

    hs = None
    hp = os.path.join(a.repo, 'HASHES.json')
    if a.update_hashes and os.path.exists(hp) and not a.dry_run:
        print(f"[3/5] HASHES.json")
        hs = sync_hashes_json(hp, a.excel, sc, ns)
        print(f"  {hs['total']} fichiers | gen_code_hash:{hs['gen_code_hash']} | drifts:{hs['drifts']}")
    else: print("[3/5] skip")

    print("[4/5] Rapport")
    rpt = report(es, hs, sc, ns)
    if a.report and not a.dry_run:
        with open(a.report,'w',encoding='utf-8') as f: f.write(rpt)
        print(f"  -> {a.report}")
    else: print(rpt)

    print(f"\n[5/5] OK")

if __name__=='__main__': sys.exit(main())
